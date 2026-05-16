#!/usr/bin/env python3
"""
vocab_converter.py
Reads china_lexicon.xlsx and writes vocab.json for the China Vocab PWA.

Run this whenever the xlsx is updated (e.g. after a new newsletter issue):
    cd ~/Documents/Claude/Projects/china-vocab-app
    python3 vocab_converter.py

The script auto-detects the xlsx path relative to this script's location.
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl

# ── Paths ─────────────────────────────────────────────────────────────────────

SCRIPT_DIR   = Path(__file__).parent
XLSX_PATH    = SCRIPT_DIR.parent / "China newsletter" / "china_lexicon.xlsx"
OUTPUT_PATH  = SCRIPT_DIR / "vocab.json"

if not XLSX_PATH.exists():
    print(f"ERROR: Could not find xlsx at {XLSX_PATH}")
    print("Edit XLSX_PATH at the top of this script if your file is elsewhere.")
    sys.exit(1)

# ── Read sheets ──────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(XLSX_PATH)
vocab = []

# AI & Tech Terms
ws = wb["AI & Tech Terms"]
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[0] and row[1] and row[2]:
        vocab.append({
            "zh":         str(row[0]).strip(),
            "pinyin":     str(row[1]).strip(),
            "en":         str(row[2]).strip(),
            "category":   str(row[3]).strip() if row[3] else "AI & Tech",
            "definition": str(row[4]).strip() if row[4] else "",
            "example":    str(row[5]).strip() if row[5] else "",
            "date_added": str(row[6]).strip() if row[6] else "",
            "source":     str(row[7]).strip() if row[7] else "",
            "type":       "term",
            "sheet":      "AI & Tech"
        })

# Gov & Finance Terms
ws = wb["Gov & Finance Terms"]
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[0] and row[1] and row[2]:
        vocab.append({
            "zh":         str(row[0]).strip(),
            "pinyin":     str(row[1]).strip(),
            "en":         str(row[2]).strip(),
            "category":   str(row[3]).strip() if row[3] else "Gov & Finance",
            "definition": str(row[4]).strip() if row[4] else "",
            "example":    str(row[5]).strip() if row[5] else "",
            "date_added": str(row[6]).strip() if row[6] else "",
            "source":     str(row[7]).strip() if row[7] else "",
            "type":       "term",
            "sheet":      "Gov & Finance"
        })

# Key People
ws = wb["Key People"]
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[0] and row[1] and row[2]:
        vocab.append({
            "zh":           str(row[0]).strip(),
            "pinyin":       str(row[1]).strip(),
            "en":           str(row[2]).strip(),
            "category":     "Key People",
            "definition":   str(row[5]).strip() if row[5] else "",
            "example":      "",
            "date_added":   str(row[6]).strip() if row[6] else "",
            "source":       "",
            "type":         "person",
            "sheet":        "Key People",
            "org":          str(row[4]).strip() if row[4] else "",
            "role":         str(row[3]).strip() if row[3] else "",
            "significance": str(row[5]).strip() if row[5] else ""
        })

# ── Write output ──────────────────────────────────────────────────────────────

with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump(vocab, f, ensure_ascii=False, indent=2)

print(f"✓ Wrote {len(vocab)} vocab items to {OUTPUT_PATH}")
print(f"  AI & Tech:    {sum(1 for v in vocab if v['sheet'] == 'AI & Tech')}")
print(f"  Gov & Finance:{sum(1 for v in vocab if v['sheet'] == 'Gov & Finance')}")
print(f"  Key People:   {sum(1 for v in vocab if v['sheet'] == 'Key People')}")
